import json
from io import BytesIO
from os.path import splitext

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, flt, getdate
from frappe.utils.file_manager import save_file
from frappe.utils.xlsxutils import make_xlsx
from openpyxl import load_workbook

ALLOWED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_WORKBOOK_FILE_SIZE_BYTES = 5 * 1024 * 1024
SYNC_ROW_LIMIT = 200
REQUIRED_COLUMNS = {"employee", "salary structure"}


class SalaryStructureAssignmentImport(Document):
	pass


@frappe.whitelist()
def import_workbook(doc_name):
	doc = frappe.get_doc("Salary Structure Assignment Import", doc_name)
	frappe.has_permission(doc.doctype, "write", doc=doc, throw=True)

	if not doc.workbook:
		frappe.throw(_("Please attach a workbook first."))

	rows = _read_workbook_rows(doc.workbook)
	if not rows:
		frappe.throw(_("No data rows found in the workbook."))

	if len(rows) > SYNC_ROW_LIMIT:
		doc.db_set({"status": "Queued", "total_rows": len(rows)}, commit=True)
		frappe.enqueue(_process_workbook_rows, queue="long", timeout=3000, doc_name=doc.name)
		return {"queued": True, "total_rows": len(rows)}

	return _process_workbook_rows(doc_name=doc.name)


@frappe.whitelist()
def retry_failed_rows(doc_name):
	"""Re-process only the rows that previously failed, merging results back into the log."""
	doc = frappe.get_doc("Salary Structure Assignment Import", doc_name)
	frappe.has_permission(doc.doctype, "write", doc=doc, throw=True)

	if not doc.import_log:
		frappe.throw(_("No import log found. Run a full import first."))
	try:
		existing_log = json.loads(doc.import_log)
	except Exception:
		frappe.throw(_("Import log is corrupted. Re-run the full import."))

	only_rows = {r["row"] for r in existing_log if r.get("status") == "Failed"}
	if not only_rows:
		frappe.throw(_("No failed rows to retry."))

	return _process_workbook_rows(doc_name=doc_name, only_rows=only_rows, existing_log=existing_log)


@frappe.whitelist()
def cancel_import(doc_name):
	"""Cancel all submitted Salary Structure Assignments created by this import."""
	frappe.has_permission("Salary Structure Assignment Import", "write", throw=True)
	frappe.has_permission("Salary Structure Assignment", "cancel", throw=True)

	ssa_names = frappe.get_all(
		"Salary Structure Assignment",
		filters={"custom_import_reference": doc_name, "docstatus": 1},
		pluck="name",
	)
	if not ssa_names:
		frappe.throw(_("No active Salary Structure Assignments are linked to this import."))

	cancelled = 0
	errors = []
	for name in ssa_names:
		try:
			ssa_doc = frappe.get_doc("Salary Structure Assignment", name)
			ssa_doc.cancel()
			cancelled += 1
		except Exception as e:
			errors.append(f"{name}: {cstr(e)}")

	new_status = "Cancelled" if not errors else "Cancelled with Errors"
	frappe.db.set_value(
		"Salary Structure Assignment Import", doc_name, "status", new_status, update_modified=False
	)
	frappe.db.commit()

	return {"cancelled": cancelled, "errors": errors, "status": new_status}


def _process_workbook_rows(doc_name, only_rows=None, existing_log=None):
	from hrms.payroll.doctype.salary_structure.salary_structure import (
		create_salary_structure_assignment,
	)
	from hr_suite.hr_suite.doctype.salary_breakup_table.salary_breakup_table import (
		get_breakup_for_total_salary,
	)

	doc = frappe.get_doc("Salary Structure Assignment Import", doc_name)
	rows = _read_workbook_rows(doc.workbook)

	results = []
	success = skipped = failed = 0

	for idx, row in enumerate(rows, start=2):  # row 1 is the header
		if only_rows and idx not in only_rows:
			continue
		employee_value = row.get("employee")
		structure_value = row.get("salary_structure")
		from_date = row.get("from_date") or doc.default_from_date
		base = row.get("base")
		variable = row.get("variable")
		total_salary = row.get("total_salary")

		row_result = {
			"row": idx,
			"employee": employee_value,
			"salary_structure": structure_value,
			"total_salary": total_salary,
			"breakup_band": None,
		}

		if not employee_value or not structure_value:
			failed += 1
			row_result.update(status="Failed", message=_("Employee and Salary Structure are both required."))
			results.append(row_result)
			continue

		if not from_date:
			failed += 1
			row_result.update(
				status="Failed",
				message=_("From Date is required (set it on the row or as the Default From Date)."),
			)
			results.append(row_result)
			continue

		savepoint = f"ssai_row_{idx}"
		try:
			frappe.db.savepoint(savepoint)

			employee_id, employee_company = _resolve_employee(employee_value, doc.company)
			structure = _resolve_salary_structure(structure_value)

			existing_ssa = frappe.db.get_value(
				"Salary Structure Assignment",
				{
					"employee": employee_id,
					"salary_structure": structure.name,
					"from_date": getdate(from_date),
					"docstatus": 1,
				},
				"name",
			)

			if existing_ssa:
				# SSA already exists — still apply breakup if total_salary given
				ssa_name = existing_ssa
				skipped += 1
				row_result.update(status="Skipped", message=_("Already assigned for this From Date."))
			else:
				# Resolve base: prefer breakup basic if total_salary given, else use base column
				resolved_base = flt(base) if base not in (None, "") else None
				if total_salary not in (None, ""):
					breakup = get_breakup_for_total_salary(flt(total_salary), employee_company)
					if breakup:
						resolved_base = breakup["basic"]

				ssa_name = create_salary_structure_assignment(
					employee=employee_id,
					salary_structure=structure.name,
					company=employee_company,
					currency=structure.currency,
					from_date=getdate(from_date),
					base=resolved_base,
					variable=flt(variable) if variable not in (None, "") else None,
				)
				# Tag the new SSA with this import record for traceability
				if frappe.db.has_column("Salary Structure Assignment", "custom_import_reference"):
					frappe.db.set_value(
						"Salary Structure Assignment", ssa_name,
						"custom_import_reference", doc_name, update_modified=False
					)
				success += 1
				row_result.update(status="Assigned", message="", ssa_name=ssa_name)

			# Apply breakup custom fields if total_salary column is present
			if total_salary not in (None, "") and ssa_name:
				breakup = get_breakup_for_total_salary(flt(total_salary), employee_company)
				if breakup:
					_write_breakup_to_ssa(ssa_name, flt(total_salary), breakup, import_name=doc_name)
					row_result["breakup_band"] = breakup["matched_total"]
				else:
					if row_result.get("status") != "Failed":
						row_result["message"] = (row_result.get("message") or "") + _(
							" Warning: no breakup band found for Total Salary {0} for company {1}."
						).format(total_salary, employee_company)

		except Exception as e:
			frappe.db.rollback(save_point=savepoint)
			failed += 1
			row_result.update(status="Failed", message=cstr(e))
			frappe.log_error(
				title="Salary Structure Assignment Import Row Failed",
				message=frappe.get_traceback(),
			)
		results.append(row_result)

	# When retrying: merge new results back into the existing log and recalculate totals
	if only_rows and existing_log:
		by_row = {r["row"]: r for r in existing_log}
		for r in results:
			by_row[r["row"]] = r
		results = sorted(by_row.values(), key=lambda r: r["row"])
		success  = sum(1 for r in results if r.get("status") == "Assigned")
		skipped  = sum(1 for r in results if r.get("status") == "Skipped")
		failed   = sum(1 for r in results if r.get("status") == "Failed")

	doc.db_set(
		{
			"total_rows": len(results) if only_rows else len(rows),
			"success_count": success,
			"skipped_count": skipped,
			"failed_count": failed,
			"status": "Completed" if failed == 0 else "Completed with Errors",
			"import_log": json.dumps(results, indent=2, default=str),
		},
		commit=True,
	)

	summary = {
		"total_rows": len(results) if only_rows else len(rows),
		"success_count": success,
		"skipped_count": skipped,
		"failed_count": failed,
		"results": results,
	}
	frappe.publish_realtime("salary_structure_assignment_import_done", summary, user=frappe.session.user)
	return summary


def _write_breakup_to_ssa(ssa_name, total_salary, breakup, import_name=None):
	"""Write breakup amounts directly to SSA custom fields (fast path for bulk import)."""
	fields = {
		"custom_total_salary": total_salary,
		"base": breakup["basic"],
		"custom_hra_amount": breakup["hra"],
		"custom_transport_amount": breakup["transport"],
		"custom_other_allowance_amount": breakup["other_allowance"],
	}
	if import_name and frappe.db.has_column("Salary Structure Assignment", "custom_import_reference"):
		fields["custom_import_reference"] = import_name
	for fieldname, value in fields.items():
		if frappe.db.has_column("Salary Structure Assignment", fieldname):
			frappe.db.set_value(
				"Salary Structure Assignment", ssa_name, fieldname, value, update_modified=False
			)


def _resolve_employee(value, default_company):
	value = cstr(value).strip()
	employee_id = value if frappe.db.exists("Employee", value) else None

	if not employee_id:
		employee_id = frappe.db.get_value("Employee", {"user_id": value}, "name")
	if not employee_id:
		# employee_name is not unique — only use if exactly one match
		matches = frappe.get_all("Employee", filters={"employee_name": value}, pluck="name", limit=2)
		if len(matches) == 1:
			employee_id = matches[0]
		elif len(matches) > 1:
			frappe.throw(
				_("Multiple employees found with name {0}. Use the Employee ID instead.").format(value)
			)
	if not employee_id:
		frappe.throw(_("No Employee found matching {0}.").format(value))

	company = frappe.db.get_value("Employee", employee_id, "company") or default_company
	if not company:
		frappe.throw(_("Employee {0} has no Company set.").format(employee_id))
	return employee_id, company


def _resolve_salary_structure(value):
	value = cstr(value).strip()
	structure = frappe.db.get_value(
		"Salary Structure", value, ["name", "currency", "docstatus"], as_dict=True
	)
	if not structure:
		frappe.throw(_("No Salary Structure found named {0}.").format(value))
	if structure.docstatus != 1:
		frappe.throw(_("Salary Structure {0} is not submitted.").format(value))
	return structure


def _read_workbook_rows(file_url):
	file_row = frappe.db.get_value(
		"File", {"file_url": file_url}, ["name", "file_name", "file_size"], as_dict=True
	)
	if not file_row:
		frappe.throw(_("Unable to find the uploaded workbook."))

	extension = splitext(cstr(file_row.file_name or file_url).strip())[1].lower()
	if extension not in ALLOWED_WORKBOOK_EXTENSIONS:
		frappe.throw(_("Only .xlsx or .xlsm files are supported."), title=_("Invalid File Type"))
	if flt(file_row.file_size) > MAX_WORKBOOK_FILE_SIZE_BYTES:
		frappe.throw(
			_("The uploaded workbook is too large. Please keep it under {0} MB.").format(
				MAX_WORKBOOK_FILE_SIZE_BYTES // (1024 * 1024)
			),
			title=_("Workbook Too Large"),
		)

	content = frappe.get_doc("File", file_row.name).get_content()
	workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
	sheet = workbook.active

	header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
	if not header_cells:
		frappe.throw(_("The workbook has no header row."))

	header_map = {}
	for i, cell in enumerate(header_cells):
		label = cstr(cell.value).strip().lower()
		if label:
			header_map[label] = i

	missing = REQUIRED_COLUMNS - set(header_map)
	if missing:
		frappe.throw(
			_("The workbook is missing required column(s): {0}.").format(", ".join(sorted(missing))),
			title=_("Missing Columns"),
		)

	def get(values, label):
		i = header_map.get(label)
		return values[i] if i is not None and i < len(values) else None

	rows = []
	for values in sheet.iter_rows(min_row=2, values_only=True):
		if not values or all(v in (None, "") for v in values):
			continue
		rows.append(
			{
				"employee": get(values, "employee"),
				"salary_structure": get(values, "salary structure"),
				"from_date": get(values, "from date"),
				"base": get(values, "base"),
				"variable": get(values, "variable"),
				"total_salary": get(values, "total salary"),
			}
		)
	return rows


@frappe.whitelist()
def download_template(doc_name):
	doc = frappe.get_doc("Salary Structure Assignment Import", doc_name)
	frappe.has_permission(doc.doctype, "read", doc=doc, throw=True)

	filters = {"status": "Active"}
	if doc.company:
		filters["company"] = doc.company

	employees = frappe.get_all(
		"Employee",
		filters=filters,
		fields=["name", "employee_name"],
		order_by="employee_name asc",
	)

	rows = [["Employee", "Employee Name", "Salary Structure", "From Date", "Total Salary", "Base", "Variable"]]
	for emp in employees:
		rows.append([emp.name, emp.employee_name, "", "", "", "", ""])

	file_name = f"salary-structure-assignment-template-{doc.name}.xlsx"
	file_doc = save_file(
		file_name,
		make_xlsx(rows, "Salary Structure Assignment").getvalue(),
		doc.doctype,
		doc.name,
		is_private=1,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name, "row_count": len(rows) - 1}
