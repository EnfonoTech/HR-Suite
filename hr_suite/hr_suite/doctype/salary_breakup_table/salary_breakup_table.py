from io import BytesIO
from os.path import splitext

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, flt, now_datetime
from openpyxl import load_workbook

from hr_suite.hr_suite.utils import assert_doctype_permissions

ALLOWED_IMPORT_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_HEADER_SCAN_ROWS = 5

# Components created/used by the breakup structure — in order of appearance on the slip
# (component_name, default_abbr, ssa_formula, is_tax_applicable)
BREAKUP_COMPONENTS = [
    ("Basic",                    "B",   "base",                        0),
    ("HRA / Living Allowances",  "HRA", "custom_hra_amount",           0),
    ("Transport / Food Allowance","TFA","custom_transport_amount",      0),
    ("Other Allowance",          "OA",  "custom_other_allowance_amount",0),
]


class SalaryBreakupTable(Document):
    pass


def _classify_header(label):
    label = cstr(label).strip().lower()
    if not label:
        return None
    if "total" in label and "salary" in label:
        return "total_salary"
    if "total" in label and ("basic" in label or "allowance" in label):
        # The sheet's own "Total Basic & Allowances" check column — not a split value.
        return None
    if "basic" in label:
        return "basic"
    if "hra" in label or "living" in label:
        return "hra"
    if "transport" in label or "food" in label:
        return "transport"
    if "other" in label:
        return "other_allowance"
    return None


def _get_file_bytes(file_url):
    file_row = frappe.db.get_value(
        "File", {"file_url": file_url}, ["name", "file_name", "file_size"], as_dict=True
    )
    if not file_row:
        frappe.throw(_("Unable to find the uploaded workbook."))
    extension = splitext(cstr(file_row.file_name or file_url).strip())[1].lower()
    if extension not in ALLOWED_IMPORT_EXTENSIONS:
        frappe.throw(_("Only Excel workbook files are supported."), title=_("Invalid File Type"))
    if flt(file_row.file_size) > MAX_IMPORT_FILE_SIZE_BYTES:
        frappe.throw(
            _("The uploaded workbook is too large. Please keep it under {0} MB.").format(
                MAX_IMPORT_FILE_SIZE_BYTES // (1024 * 1024)
            ),
            title=_("Workbook Too Large"),
        )
    return frappe.get_doc("File", file_row.name).get_content()


def _find_header_row(sheet):
    """Scan the first few rows for the real header (skips merged title rows like 'Salary Breakup')."""
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS), start=1
    ):
        header_map = {}
        for col_idx, cell in enumerate(row):
            key = _classify_header(cell.value)
            if key and key not in header_map:
                header_map[key] = col_idx
        if {"total_salary", "basic"}.issubset(header_map):
            return row_idx, header_map
    return None, {}


def _read_breakup_rows(file_url):
    content = _get_file_bytes(file_url)
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active

    header_row_idx, header_map = _find_header_row(sheet)
    if header_row_idx is None:
        frappe.throw(
            _("Could not find a header row with Total Salary and Basic columns in the workbook."),
            title=_("Missing Columns"),
        )

    def get(values, key):
        i = header_map.get(key)
        return values[i] if i is not None and i < len(values) else None

    rows = []
    for values in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not values or all(v in (None, "") for v in values):
            continue
        total_salary = get(values, "total_salary")
        if total_salary in (None, ""):
            continue
        rows.append(
            {
                "total_salary": flt(total_salary),
                "basic": flt(get(values, "basic")),
                "hra": flt(get(values, "hra")),
                "transport": flt(get(values, "transport")),
                "other_allowance": flt(get(values, "other_allowance")),
            }
        )
    return rows


@frappe.whitelist()
def import_breakup_table(doc_name, file_url=None):
    """Import breakup rows into the company-specific Salary Breakup Table record.

    doc_name is the company name (record is named after the company).
    """
    assert_doctype_permissions("Salary Breakup Table", "write")

    if frappe.db.exists("Salary Breakup Table", doc_name):
        doc = frappe.get_doc("Salary Breakup Table", doc_name)
    else:
        doc = frappe.new_doc("Salary Breakup Table")
        doc.company = doc_name

    file_url = file_url or doc.breakup_workbook
    if not file_url:
        frappe.throw(_("Attach a Salary Breakup Workbook first."))

    rows = _read_breakup_rows(file_url)
    if not rows:
        frappe.throw(_("No usable rows found in the workbook."))

    doc.breakup_workbook = file_url
    doc.set("rows", [])
    for row in rows:
        doc.append("rows", row)
    doc.last_imported_on = now_datetime()
    doc.imported_row_count = len(rows)
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"row_count": len(rows)}


def get_breakup_for_total_salary(total_salary, company):
    """Return the breakup for the nearest table band at or below total_salary.

    Exact match wins; otherwise falls back to the highest band ≤ the requested
    amount so real salaries (not divisible by 100) still resolve correctly.
    Returns None if no record exists for the company or the salary is below all bands.
    """
    total_salary = flt(total_salary)
    if not company or not frappe.db.exists("Salary Breakup Table", company):
        return None

    doc = frappe.get_doc("Salary Breakup Table", company)
    best = None
    best_val = -1.0

    for row in doc.rows or []:
        row_val = flt(row.total_salary)
        if row_val == total_salary:
            return {
                "matched_total": row_val,
                "basic": flt(row.basic),
                "hra": flt(row.hra),
                "transport": flt(row.transport),
                "other_allowance": flt(row.other_allowance),
            }
        if row_val < total_salary and row_val > best_val:
            best_val = row_val
            best = row

    if best:
        return {
            "matched_total": best_val,
            "basic": flt(best.basic),
            "hra": flt(best.hra),
            "transport": flt(best.transport),
            "other_allowance": flt(best.other_allowance),
        }
    return None


@frappe.whitelist()
def get_breakup_preview(employee, total_salary):
    """Return the breakup split for the dialog preview.

    Derives the company directly from the employee record.
    """
    frappe.has_permission("Salary Breakup Table", "read", throw=True)
    company = frappe.db.get_value("Employee", employee, "company")
    return get_breakup_for_total_salary(flt(total_salary), company)


@frappe.whitelist()
def create_salary_structure_from_breakup(company):
    """Create a draft monthly Salary Structure for the given company.

    Creates any missing Salary Component records first, then builds the structure
    with formula-based earnings that read from SSA custom fields populated by
    apply_salary_breakup.  Leaves the structure in draft so HR can review
    taxability settings before submitting.
    """
    frappe.has_permission("Salary Structure", "create", throw=True)

    struct_name = f"{company} Common Structure"
    if frappe.db.exists("Salary Structure", struct_name):
        frappe.throw(
            _("Salary Structure {0} already exists. Open it from Payroll → Salary Structure.").format(
                struct_name
            )
        )

    currency = (
        frappe.db.get_value("Company", company, "default_currency")
        or frappe.db.get_default("currency")
        or "USD"
    )

    # Ensure each Salary Component exists; create if not
    for comp_name, default_abbr, _formula, is_tax in BREAKUP_COMPONENTS:
        if not frappe.db.exists("Salary Component", comp_name):
            frappe.get_doc(
                {
                    "doctype": "Salary Component",
                    "salary_component": comp_name,
                    "salary_component_abbr": default_abbr,
                    "type": "Earning",
                    "is_tax_applicable": is_tax,
                }
            ).insert(ignore_permissions=True)

    # Build earnings rows — use the abbr already on the component in case it exists
    earnings = []
    for comp_name, default_abbr, formula, _ in BREAKUP_COMPONENTS:
        abbr = (
            frappe.db.get_value("Salary Component", comp_name, "salary_component_abbr")
            or default_abbr
        )
        earnings.append(
            {
                "salary_component": comp_name,
                "abbr": abbr,
                "amount_based_on_formula": 1,
                "formula": formula,
                "amount": 0,
                "depends_on_payment_days": 1,
            }
        )

    structure = frappe.get_doc(
        {
            "doctype": "Salary Structure",
            "name": struct_name,
            "payroll_frequency": "Monthly",
            "is_active": "Yes",
            "currency": currency,
            "earnings": earnings,
            "deductions": [],
        }
    )
    structure.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "name": structure.name,
        "currency": currency,
        "components": [c[0] for c in BREAKUP_COMPONENTS],
    }
